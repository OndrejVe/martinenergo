import os
import json
import random
import re
import logging
import unicodedata
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

import boto3
import numpy as np
try:
    from openai import OpenAI  # type: ignore
except ImportError:  # SDK nemusí být v lokálním prostředí
    OpenAI = None  # type: ignore
from openpyxl import load_workbook

from backend.services.tdd_prices import get_yearly_tdd_prices

RAG_BUCKET = os.getenv("RAG_BUCKET", "")
RAG_PREFIX = os.getenv("RAG_PREFIX", "index/")
TOP_K = int(os.getenv("TOP_K", "5"))

REG = os.getenv("AWS_REGION", "eu-central-1")
EMB_ID = os.getenv("EMBEDDINGS_MODEL_ID", "amazon.titan-embed-text-v2:0")
CHAT_ID = os.getenv("CHAT_MODEL_ID", "amazon.titan-text-lite-v1")
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
ENABLE_BEDROCK = os.getenv("ENABLE_BEDROCK", "1") not in ("0", "false", "False")

logger = logging.getLogger(__name__)

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = PROJECT_ROOT / "rag" / "docs" / "data"

br = None
s3 = boto3.client("s3")
_OPENAI = None

INDEX_LOCAL = "/tmp/index.npz"
CACHE = {"V": None, "chunks": None}
LEX = {"matrix": None, "idf": None, "vocab": None}
SAZBA_TO_TDD: Dict[str, str] = {}
TDD_PRICES: Dict[str, float] = {}

DEFAULT_SAZBA = "D25D"
DEFAULT_TDD = "TDD5"
DEFAULT_CONSUMPTION_MWH = 3.0
FIX_MARKUP = 0.12

LEAD_ANALYZE_LINE = (
    "Dobrý den, rádi bychom Vám nabídli analýzu spotřeby a výpočty úspor, které by mohly pomoci optimalizovat náklady na energie ve vaší firmě."
)
LEAD_DIFFERENCE_LINE = (
    "Na základě historických dat o spotřebě energie vaší firmy Vám mohu spočítat rozdíl mezi fixním a spotovým tarifem a ukázat, jak to ovlivní roční náklady."
)
LEAD_GUIDE_LINE = (
    "Pokud máte zájem, rádi Vám ukážeme, kde ve Vašem vyúčtování najdete informace potřebné pro výpočet úspor a připravíme srovnání vhodných tarifů."
)
LEAD_CONSUMPTION_QUESTION = (
    "Můžete mi prosím poskytnout odhad roční spotřeby vaší firmy v MWh? Díky tomu zpřesním výpočet úspor."
)
LEAD_SPOT_QUESTION = (
    "Chcete zjistit, jaké by byly úspory při přechodu na spotový tarif? Mohu vám připravit konkrétní scénář."
)
EMAIL_REQUEST_PROMPT = (
    "Pokud chcete, rádi vám zašleme na e-mail detailní nabídku ve formátu PDF – stačí mi poslat adresu, kam ji můžeme doručit."
)

LEAD_LINES_BUSINESS = (
    LEAD_ANALYZE_LINE,
    LEAD_DIFFERENCE_LINE,
    LEAD_GUIDE_LINE,
    LEAD_CONSUMPTION_QUESTION,
    LEAD_SPOT_QUESTION,
    EMAIL_REQUEST_PROMPT,
)

SAVINGS_KEYWORDS = (
    "uspora",
    "uspor",
    "usetrit",
    "spotov",
    "fixni tarif",
    "tarif",
    "cenik",
    "vyuctovan",
    "silova elektrina",
    "cena silove elektriny",
    "cena elektřiny",
    "cena elektriny",
    "cena za silovou elektřinu",
    "cena za silovou elektrinu",
    "kde najdu cenu",
    "kde najdu cenu elektřiny",
    "kde najdu cenu elektriny",
    "cena na faktuře",
    "cena na vyúčtování",
    "cena na vyuctovani",
    "silová elektřina",
)
SILOVA_ELEKTRINA_EXPLANATION = (
    "Silová elektřina je samotná komodita, tedy elektrická energie, kterou skutečně spotřebujete. "
    "Je to hlavní položka na vaší faktuře za elektřinu, která tvoří podstatnou část celkové ceny. "
    "Kromě silové elektřiny platíte ještě regulované poplatky (za distribuci, systémové služby apod.), které stanovuje stát. "
    "Cena silové elektřiny je určena vaším dodavatelem a může být buď fixní (pevně sjednaná na určité období), nebo spotová (odvíjí se od aktuální ceny na burze). "
    "\n\n"
    "Jak najít cenu silové elektřiny na faktuře?\n"
    "Na faktuře ji obvykle najdete v části s názvem 'Silová elektřina' nebo 'Dodávka elektřiny'. "
    "Bývá uvedena v Kč/MWh nebo Kč/kWh. Pokud si nejste jistí, kde přesně tuto částku najít, rád vám s tím pomohu – stačí mi poslat kopii faktury, nebo část s vyúčtováním.\n"
    "\n"
    "Pokud chcete poradit s výběrem vhodného tarifu nebo porovnat ceny, stačí mi poskytnout vaši roční spotřebu a typ sazby – připravím vám konkrétní srovnání."
)
HOUSEHOLD_KEYWORDS = ("domacnost", "byt", "rodina", "bytovy", "domaci")

HOUSEHOLD_NOTICE = (
    "Tato služba je výhradně určena pro firemní klienty. "
    "Rádi Vám však spočítáme úspory a poskytneme výpočet i pro domácnosti. "
    "Pokud budete mít zájem, rádi vám ukážeme, jaký rozdíl mezi fixním a spotovým tarifem by pro Vás mohl být nejvýhodnější."
)

COMPETITION_KEYWORDS = ("konkurenc", "jiny dodavatel", "jineho dodavatele", "cez", "eon", "innogy", "bohemia energy", "pražská energetika", "pre")
WEATHER_KEYWORDS = ("pocasi", "teplota", "venku", "predpoved", "meteo")
TESTING_KEYWORDS = ("test", "zkouska", "zkousim", "hraju", "haha", "lol", "nesmysl")

COMPETITION_NOTICE = "V tomto směru vám nemohu poskytnout odpověď, ale rád vám pomohu zjistit, jaké možnosti úspor máme my!"
WEATHER_NOTICE = "Rád vám pomohu zjistit, jak můžete ušetřit na energiích, ale pokud jde o počasí, doporučuji se podívat na meteorologické weby."
TESTING_NOTICE = (
    "Pokud máte konkrétní otázky týkající se úspor nebo energetických tarifů, rád je zodpovím. "
    "Jinak vám mohu poskytnout informace o tom, jak naši službu využít nejefektivněji pro vaši firmu, případně vás kontaktovat e-mailem."
)

EMAIL_ACK_TEMPLATE = (
    "Skvělé! Na adresu {email} vám během chvíle zašleme PDF s výpočtem úspor a doporučením tarifů. "
    "Kdybyste potřebovali něco upřesnit, dejte mi vědět."
)

EMAIL_REGEX = re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")


def _get_bedrock():
    global br
    if not ENABLE_BEDROCK:
        return None
    if br is False:
        return None
    if br is None:
        try:
            session = boto3.session.Session()
            if session.get_credentials() is None:
                logger.warning("Bedrock klient není dostupný – chybí AWS pověření.")
                br = False
                return None
            br = session.client("bedrock-runtime", region_name=REG)
        except Exception as exc:  # pragma: no cover - pouze při špatných pověřeních
            logger.warning("Bedrock klient nelze inicializovat: %s", exc)
            br = False
            return None
    return br


def _ensure_index():
    if CACHE["chunks"] is not None:
        return
    local = PROJECT_ROOT / "rag" / "out" / "index.npz"
    path = local if local.exists() else None
    if path is None:
        s3.download_file(RAG_BUCKET, RAG_PREFIX + "index.npz", INDEX_LOCAL)
        path = Path(INDEX_LOCAL)
    data = np.load(path, allow_pickle=True)
    V = data["vectors"].astype("float32")
    if V.size:
        V /= (np.linalg.norm(V, axis=1, keepdims=True) + 1e-9)
        CACHE["V"] = V
    CACHE["chunks"] = data["chunks"].tolist()
    _build_lex_index()


def _build_lex_index():
    if LEX["matrix"] is not None:
        return
    texts = CACHE.get("chunks") or []
    tokens = [re.findall(r"\w+", (t or "").lower()) for t in texts]
    vocab: Dict[str, int] = {}
    for tok_list in tokens:
        for tok in tok_list:
            if tok not in vocab:
                vocab[tok] = len(vocab)
    if not vocab:
        LEX["matrix"] = np.zeros((len(texts), 0), dtype="float32")
        LEX["idf"] = np.zeros((0,), dtype="float32")
        LEX["vocab"] = {}
        return
    df = np.zeros(len(vocab), dtype="float32")
    for tok_list in tokens:
        for tok in set(tok_list):
            df[vocab[tok]] += 1
    idf = np.log((1 + len(texts)) / (1 + df)) + 1
    matrix = np.zeros((len(texts), len(vocab)), dtype="float32")
    for i, tok_list in enumerate(tokens):
        if not tok_list:
            continue
        counts = Counter(tok_list)
        denom = float(len(tok_list))
        for tok, cnt in counts.items():
            j = vocab[tok]
            matrix[i, j] = (cnt / denom) * idf[j]
        norm = np.linalg.norm(matrix[i]) + 1e-9
        matrix[i] /= norm
    LEX["matrix"] = matrix
    LEX["idf"] = idf
    LEX["vocab"] = vocab


def _ensure_tariff_assets():
    global SAZBA_TO_TDD, TDD_PRICES
    if not SAZBA_TO_TDD:
        try:
            wb = load_workbook(DATA_DIR / "D_sazba_TDD vazby.xlsx", data_only=True)
            ws = wb.active
            for sazba, tdd in ws.iter_rows(min_row=2, values_only=True):
                if sazba and tdd:
                    SAZBA_TO_TDD[sazba.strip().upper()] = tdd.strip().upper()
        except Exception as exc:
            logger.warning("Nepodařilo se načíst mapování sazeb: %s", exc)
    if not TDD_PRICES:
        try:
            yearly_prices = get_yearly_tdd_prices()
            for tdd, price in yearly_prices.items():
                if price:
                    TDD_PRICES[tdd.upper()] = float(price)
            if DEFAULT_TDD not in TDD_PRICES and TDD_PRICES:
                TDD_PRICES[DEFAULT_TDD] = sum(TDD_PRICES.values()) / len(TDD_PRICES)
        except FileNotFoundError as exc:
            logger.warning("Chybí Excel s reálnými TDD cenami: %s", exc)
        except Exception as exc:
            logger.warning("Nepodařilo se načíst modelové ceny z Excelu: %s", exc)
    if not TDD_PRICES:
        TDD_PRICES[DEFAULT_TDD] = 2700.0


def _embed_bedrock(text: str):
    client = _get_bedrock()
    if client is None:
        raise RuntimeError("Bedrock není k dispozici.")
    body = json.dumps({"inputText": text})
    r = client.invoke_model(modelId=EMB_ID, body=body)
    v = np.array(json.loads(r["body"].read())["embedding"], dtype="float32")
    v /= (np.linalg.norm(v) + 1e-9)
    return v


def _lexical_vector(text: str):
    vocab = LEX.get("vocab") or {}
    if not vocab:
        return np.zeros((0,), dtype="float32")
    tokens = re.findall(r"\w+", text.lower())
    if not tokens:
        return np.zeros((len(vocab),), dtype="float32")
    counts = Counter(tokens)
    vec = np.zeros((len(vocab),), dtype="float32")
    idf = LEX["idf"]
    denom = float(len(tokens))
    for tok, cnt in counts.items():
        j = vocab.get(tok)
        if j is None:
            continue
        vec[j] = (cnt / denom) * idf[j]
    norm = np.linalg.norm(vec) + 1e-9
    return vec / norm


def _retrieve_from_matrix(matrix: np.ndarray, qv: np.ndarray, k: int) -> List[dict]:
    if matrix is None or qv.size == 0 or matrix.shape[1] != qv.shape[0]:
        return []
    k = min(k, matrix.shape[0])
    sims = matrix @ qv
    idx = np.argpartition(-sims, k - 1)[:k]
    idx = idx[np.argsort(-sims[idx])]
    return [{"score": float(sims[i]), "text": CACHE["chunks"][int(i)][:2000]} for i in idx]


def _retrieve_hits(query: str, k: int) -> List[dict]:
    _ensure_index()
    V = CACHE.get("V")
    if V is not None:
        try:
            qv = _embed_bedrock(query)
            return _retrieve_from_matrix(V, qv, k)
        except Exception as exc:
            logger.warning("Vektorové vyhledávání přes Bedrock selhalo (%s), přepínám na TF-IDF.", exc)
    qv = _lexical_vector(query)
    matrix = LEX["matrix"]
    if matrix is None:
        return []
    k = min(k, matrix.shape[0])
    sims = matrix @ qv if qv.size else np.zeros(matrix.shape[0])
    idx = np.argpartition(-sims, k - 1)[:k]
    idx = idx[np.argsort(-sims[idx])]
    return [{"score": float(sims[i]), "text": CACHE["chunks"][int(i)][:2000]} for i in idx]


def _fallback_answer(hits: List[dict]) -> str:
    if not hits:
        return _append_lead_hint("Omlouvám se, ale v tomto režimu nemám k dispozici žádná data pro odpověď.")
    snippet = hits[0]["text"].strip()
    return _append_lead_hint("Lokální režim (bez LLM):\n" + snippet[:500])


def _append_lead_hint(text: str) -> str:
    if not text.strip():
        return random.choice(LEAD_LINES_BUSINESS)
    if any(hint in text for hint in LEAD_LINES_BUSINESS):
        return text
    return text.rstrip() + "\n\n" + random.choice(LEAD_LINES_BUSINESS)


def _chat(ctx: str, q: str, hits: List[dict]) -> str:
    lead_hint = random.choice(LEAD_LINES_BUSINESS)
    prompt = (
        "Jsi Energo – firemní energetický poradce jménem Martin. Soustřeď se na B2B klientelu, "
        "buď proaktivní, ale nenásilný, navrhuj analýzu spotřeby a konkrétní kroky pro optimalizaci nákladů. "
        "Pokud uživatel zmíní domácnost, zdvořile připomeň, že služba je určena primárně firmám. "
        "Neodkazuj na obecné vyhledávače ani Perplexity/Google, používej pouze kontext a vlastní znalosti. "
        f"Klidně použij věty jako: {lead_hint}.\n\n"
        f"Kontext:\n{ctx}\n\nOtázka: {q}\nOdpověď:"
    )
    client = _get_bedrock()
    if CHAT_ID.startswith("anthropic."):
        if client is None:
            logger.warning("Anthropic model není dostupný bez Bedrocku, vracím fallback.")
            return _fallback_answer(hits)
        body = {
            "anthropic_version": "bedrock-2023-05-31",
            "max_tokens": 400,
            "messages": [{"role": "user", "content": [{"type": "text", "text": prompt}]}],
        }
        r = client.invoke_model(modelId=CHAT_ID, body=json.dumps(body))
        out = json.loads(r["body"].read())
        return _append_lead_hint(out["content"][0]["text"])
    if CHAT_ID.startswith("amazon.titan-text"):
        if client is None:
            logger.warning("Titání model není dostupný bez Bedrocku, vracím fallback.")
            return _fallback_answer(hits)
        body = {
            "inputText": prompt,
            "textGenerationConfig": {"maxTokenCount": 400, "temperature": 0.2, "topP": 0.9},
        }
        r = client.invoke_model(modelId=CHAT_ID, body=json.dumps(body))
        out = json.loads(r["body"].read())
        return _append_lead_hint(out["results"][0]["outputText"].strip())
    if CHAT_ID.startswith("openai:"):
        model = CHAT_ID.split(":", 1)[1] or "gpt-4o-mini"
        global _OPENAI
        if OpenAI is None:
            logger.warning("OpenAI SDK není nainstalováno. Přepínám na fallback odpověď.")
            return _fallback_answer(hits)
        if _OPENAI is None:
            if not OPENAI_API_KEY:
                logger.warning("Chybí OPENAI_API_KEY, nelze použít OpenAI chat. Přepínám na fallback.")
                return _fallback_answer(hits)
            _OPENAI = OpenAI(api_key=OPENAI_API_KEY)
        try:
            resp = _OPENAI.chat.completions.create(
                model=model,
                messages=[
                    {"role": "system", "content": "Jsi český firemní energetický poradce jménem Martin. Zaměř se na B2B klienty, navrhuj analýzy spotřeby a výpočet úspor a buď proaktivní, ale nenásilný."},
                    {"role": "user", "content": prompt},
                ],
                max_tokens=400,
                temperature=0.2,
            )
            return _append_lead_hint(resp.choices[0].message.content.strip())
        except Exception as exc:
            logger.warning("OpenAI odpověď selhala (%s). Přepínám na fallback.", exc)
            return _fallback_answer(hits)
    logger.warning("Chat model %s není podporovaný v lokálním režimu, vracím fallback.", CHAT_ID)
    return _fallback_answer(hits)


def _normalized(text: str) -> str:
    return unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode().lower()


def _is_savings_query(normalized_text: str) -> bool:
    return any(keyword in normalized_text for keyword in SAVINGS_KEYWORDS)


def _is_household_query(normalized_text: str) -> bool:
    return any(keyword in normalized_text for keyword in HOUSEHOLD_KEYWORDS)


def _contains_competition(normalized_text: str) -> bool:
    return any(keyword in normalized_text for keyword in COMPETITION_KEYWORDS)


def _contains_weather(normalized_text: str) -> bool:
    return any(keyword in normalized_text for keyword in WEATHER_KEYWORDS)


def _is_testing_query(normalized_text: str) -> bool:
    return any(keyword in normalized_text for keyword in TESTING_KEYWORDS)


def _extract_email(text: str) -> str:
    match = EMAIL_REGEX.search(text)
    return match.group(0) if match else ""


def _extract_sazba(text: str) -> str:
    match = re.search(r"d\d{2}d", text.lower())
    return match.group(0).upper() if match else ""


def _extract_consumption_mwh(text: str) -> float:
    match = re.search(r"(\d+(?:[\.,]\d+)?)\s*(kwh|mwh)", text.lower())
    if not match:
        return 0.0
    value = float(match.group(1).replace(",", "."))
    return value / 1000.0 if match.group(2) == "kwh" else value


def _fmt_currency(value: float) -> str:
    return f"{int(round(value))} Kč"


def compute_tariff_stats(sazba: str, consumption_mwh: float, fixed_price_kwh: Optional[float] = None) -> Dict[str, float]:
    _ensure_tariff_assets()
    sazba = (sazba or DEFAULT_SAZBA).upper()
    tdd = SAZBA_TO_TDD.get(sazba, DEFAULT_TDD)
    consumption = consumption_mwh if consumption_mwh > 0 else DEFAULT_CONSUMPTION_MWH
    spot_price = TDD_PRICES.get(tdd, TDD_PRICES.get(DEFAULT_TDD, 2700.0))
    if fixed_price_kwh is not None:
        fix_price = fixed_price_kwh * 1000.0
    else:
        fix_price = spot_price * (1 + FIX_MARKUP)
    spot_total = spot_price * consumption
    fix_total = fix_price * consumption
    return {
        "sazba": sazba,
        "tdd": tdd,
        "consumption_mwh": consumption,
        "spot_price_per_mwh": spot_price,
        "fix_price_per_mwh": fix_price,
        "spot_total": spot_total,
        "fix_total": fix_total,
    }


def calculate_business_savings(query: str) -> Dict:
    _ensure_tariff_assets()
    stats = compute_tariff_stats(_extract_sazba(query) or DEFAULT_SAZBA, _extract_consumption_mwh(query))
    sazba = stats["sazba"]
    tdd = stats["tdd"]
    consumption = stats["consumption_mwh"]
    spot_price = stats["spot_price_per_mwh"]
    fix_price = stats["fix_price_per_mwh"]
    spot_total = stats["spot_total"]
    fix_total = stats["fix_total"]
    savings = max(0.0, fix_total - spot_total)
    answer = (
        f"{LEAD_LINES_BUSINESS[0]} "
        f"Na základě historických dat o spotřebě (sazba {sazba}, třída {tdd}) vychází při "
        f"odběru ~{consumption:.1f} MWh spotový tarif na {_fmt_currency(spot_total)}, zatímco fixní "
        f"tarif by stál {_fmt_currency(fix_total)}. To znamená potenciální úsporu přibližně {_fmt_currency(savings)} ročně. "
        f"{LEAD_LINES_BUSINESS[1]} "
        f"{LEAD_LINES_BUSINESS[2]} "
        f"{EMAIL_REQUEST_PROMPT}"
    )
    chart = {
        "type": "bar",
        "title": f"Roční firemní náklady pro {consumption:.1f} MWh",
        "labels": ["Fixní tarif", "Spotový tarif"],
        "data": [round(fix_total, 2), round(spot_total, 2)],
        "meta": {
            "sazba": sazba,
            "tdd": tdd,
            "consumption_mwh": consumption,
            "year": datetime.now().year,
            "spot_price_per_mwh": spot_price,
            "fix_price_per_mwh": fix_price,
        },
    }
    return {"answer": answer, "sources": [], "chart": chart}


def _is_silova_elekt_query(normalized_text: str) -> bool:
    """
    Detects if the query is specifically about silová elektřina (electricity commodity price),
    electricity price on the bill, or directly asks for a calculation of the electricity bill.
    Now also detects direct requests for price-per-unit or total cost calculations.
    """
    silova_keywords = [
        "silova elektrina", "silová elektřina", "cena silove elektriny", "cena silové elektřiny",
        "cena elektřiny", "cena elektriny", "cena za silovou elektřinu", "cena za silovou elektrinu",
        "kde najdu cenu", "kde najdu cenu elektřiny", "kde najdu cenu elektriny",
        "cena na faktuře", "cena na vyúčtování", "cena na vyuctovani",
        "kolik stojí elektřina", "kolik stojí silová elektřina", "kolik platím za elektřinu",
        "výpočet ceny elektřiny", "výpočet účtu za elektřinu", "vyúčtování elektřiny"
    ]
    # Detect direct question about price per unit or bill
    pattern_price = r"(kolik\s+stoj[íi]|cena|výpočet|vyuctovani|vyúčtování|platím|platim|účtu|uctu|bill|price)"
    pattern_electricity = r"(elektrina|elektřina|silova|silová)"
    # If any strong keyword matches
    if any(k in normalized_text for k in silova_keywords):
        return True
    # If a price/cost/bill pattern is found together with electricity context
    if re.search(pattern_price, normalized_text) and re.search(pattern_electricity, normalized_text):
        return True
    return False


def lambda_handler(event, context):
    """
    Handles incoming API requests, integrates the RAG retrieval process,
    calculates energy savings, and returns comprehensive responses.
    Enhanced with logging and smarter detection of silová elektřina queries.
    """
    try:
        body = json.loads(event.get("body") or "{}")
    except Exception:
        body = {}
    q = (body.get("q") or "").strip()
    logger.info(f"lambda_handler input: {q!r}")
    if not q:
        logger.warning("Missing 'q' in request body.")
        return {
            "statusCode": 400,
            "headers": {"Content-Type": "application/json"},
            "body": json.dumps({"error": "missing q"}),
        }
    email = _extract_email(q)
    if email:
        logger.info(f"Detected email in query: {email}")
        return {
            "statusCode": 200,
            "headers": {"Content-Type": "application/json"},
            "body": json.dumps({"answer": EMAIL_ACK_TEMPLATE.format(email=email)}, ensure_ascii=False),
        }
    normalized = _normalized(q)
    # Handle household queries
    if _is_household_query(normalized):
        logger.info("Detected household query.")
        return {
            "statusCode": 200,
            "headers": {"Content-Type": "application/json"},
            "body": json.dumps({"answer": HOUSEHOLD_NOTICE}),
        }
    # Handle competition queries
    if _contains_competition(normalized):
        logger.info("Detected competition query.")
        return {
            "statusCode": 200,
            "headers": {"Content-Type": "application/json"},
            "body": json.dumps({"answer": COMPETITION_NOTICE}),
        }
    # Handle weather queries
    if _contains_weather(normalized):
        logger.info("Detected weather query.")
        return {
            "statusCode": 200,
            "headers": {"Content-Type": "application/json"},
            "body": json.dumps({"answer": WEATHER_NOTICE}),
        }
    # Handle testing queries
    if _is_testing_query(normalized):
        logger.info("Detected testing query.")
        return {
            "statusCode": 200,
            "headers": {"Content-Type": "application/json"},
            "body": json.dumps({"answer": TESTING_NOTICE}),
        }
    # Handle silová elektřina and power price on bill queries, including direct calculation
    if _is_silova_elekt_query(normalized):
        # Try to extract sazba and consumption for a concrete calculation
        sazba = _extract_sazba(q)
        consumption = _extract_consumption_mwh(q)
        logger.info(f"Silová elektřina query detected. Sazba: {sazba}, Consumption: {consumption}")
        # If both rate and consumption are present, provide a calculation
        if sazba or consumption > 0:
            stats = compute_tariff_stats(sazba, consumption)
            price_per_mwh = stats["spot_price_per_mwh"]
            total = stats["spot_total"]
            answer = (
                "Cena silové elektřiny (komodity) pro vaši sazbu "
                f"{stats['sazba']} ({stats['tdd']}) je přibližně {int(round(price_per_mwh))} Kč/MWh. "
                f"Při spotřebě {stats['consumption_mwh']:.2f} MWh by roční náklady na silovou elektřinu činily {int(round(total))} Kč. "
                "Tato cena nezahrnuje regulované poplatky (distribuce, systémové služby apod.), které stanovuje stát. "
                "Pokud chcete přesnější výpočet nebo porovnat s fixním tarifem, napište mi vaši sazbu a roční spotřebu."
            )
            logger.info(f"Returning silová elektřina calculation answer: {answer}")
            return {
                "statusCode": 200,
                "headers": {"Content-Type": "application/json"},
                "body": json.dumps({"answer": answer}, ensure_ascii=False),
            }
        # Otherwise, provide the explanation
        logger.info("Returning silová elektřina explanation.")
        return {
            "statusCode": 200,
            "headers": {"Content-Type": "application/json"},
            "body": json.dumps({"answer": SILOVA_ELEKTRINA_EXPLANATION}, ensure_ascii=False),
        }
    # Handle direct savings queries with calculation
    if _is_savings_query(normalized):
        logger.info("Detected savings query.")
        savings_payload = calculate_business_savings(q)
        logger.info(f"Savings calculation result: {savings_payload}")
        return {
            "statusCode": 200,
            "headers": {"Content-Type": "application/json"},
            "body": json.dumps(savings_payload, ensure_ascii=False),
        }
    # RAG process: retrieve relevant context from data sources
    hits = _retrieve_hits(q, TOP_K)
    ctx = "\n\n---\n".join([h["text"] for h in hits])
    # Generate answer using the chat model and context
    ans = _chat(ctx, q, hits)
    logger.info(f"RAG answer: {ans}")
    return {
        "statusCode": 200,
        "headers": {"Content-Type": "application/json"},
        "body": json.dumps({"answer": ans, "sources": []}, ensure_ascii=False),
    }
